import openpyxl
from flask import Flask, jsonify, request
from flask_cors import CORS
import os
import json
import pymysql
from werkzeug.utils import secure_filename
import xlrd

app = Flask(__name__)
# Ensure JSON responses are sent with UTF-8 encoding for Unicode characters
app.json.ensure_ascii = False
CORS(app)  # This will enable CORS for all routes

# Config upload folder
UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), 'uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Load database configuration
DB_CONFIG_PATH = os.path.join(os.path.dirname(__file__), 'db_config.json')
DB_CONFIG = {}
if os.path.exists(DB_CONFIG_PATH):
    with open(DB_CONFIG_PATH, 'r', encoding='utf-8') as f:
        DB_CONFIG = json.load(f)

def get_db_connection():
    return pymysql.connect(
        host=DB_CONFIG.get('DB_HOST', 'localhost'),
        user=DB_CONFIG.get('DB_USER', 'root'),
        password=DB_CONFIG.get('DB_PASSWORD', ''),
        database=DB_CONFIG.get('DB_NAME', 'qldt'),
        cursorclass=pymysql.cursors.DictCursor
    )

@app.route('/api/upload', methods=['POST'])
def upload_schedule():
    if 'file' not in request.files:
        return jsonify({"error": "Không có file nào được chọn."}), 400
    file = request.files['file']
    week = request.form.get('week')
    year = request.form.get('year')

    if file.filename == '':
        return jsonify({"error": "Tên file trống."}), 400
    if not week:
        return jsonify({"error": "Vui lòng chọn tuần."}), 400

    if file and (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
        # Preserve original extension
        ext = os.path.splitext(file.filename)[1].lower()
        safe_week = week.replace('/', '_').replace('\\', '_')
        path_prefix = safe_week
        if year:
            safe_year = year.replace('/', '_').replace('\\', '_')
            path_prefix = f"{safe_year}_{safe_week}"
            
        filename = f"{path_prefix}{ext}"
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        
        # Delete conflicting format
        alt_ext = '.xlsx' if ext == '.xls' else '.xls'
        alt_file_path = os.path.join(app.config['UPLOAD_FOLDER'], f"{path_prefix}{alt_ext}")
        if os.path.exists(alt_file_path):
            try:
                os.remove(alt_file_path)
            except OSError:
                pass
        
        try:
            file.save(file_path)
            return jsonify({"message": f"Tải lên thành công lịch cho {week}"}), 200
        except Exception as e:
            return jsonify({"error": f"Lỗi khi lưu file: {str(e)}"}), 500
    else:
        return jsonify({"error": "Chỉ chấp nhận file định dạng .xlsx hoặc .xls"}), 400


def _parse_xlsx(file_path):
    # Use openpyxl to get merged cells info
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # Build a dictionary of how each cell is merged
    merged_cells_info = {}
    for merged_range in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row_merge = merged_range.bounds
        for r in range(min_row, max_row_merge + 1):
            for c in range(min_col, max_col + 1):
                if r == min_row and c == min_col:
                    merged_cells_info[(r, c)] = {
                        'row_span': max_row_merge - min_row + 1,
                        'col_span': max_col - min_col + 1
                    }
                else:
                    merged_cells_info[(r, c)] = {
                        'row_span': 0,
                        'col_span': 0
                    }

    # Detect start row (look for "TT" or similar header)
    start_row = 1
    for r in range(1, min(sheet.max_row, 50) + 1):
        found_header = False
        for c in range(1, min(sheet.max_column, 20) + 1):
            val = sheet.cell(row=r, column=c).value
            if isinstance(val, str) and val.strip().upper() == 'TT':
                start_row = r
                found_header = True
                break
        if found_header:
            break
            
    # Detect actual max column in the table
    max_col = 1
    for c in range(1, sheet.max_column + 1):
        val = sheet.cell(row=start_row, column=c).value
        if val is not None and str(val).strip() != "":
            max_col = max(max_col, c)
    
    # In case the table starts with merged cells that span further
    for (r, c), info in merged_cells_info.items():
        if r == start_row and info['row_span'] > 0:
            max_col = max(max_col, c + info['col_span'] - 1)
            
    # To find max_col reliably based on data across a few header rows
    for r in range(start_row, min(start_row + 3, sheet.max_row + 1)):
        for c in range(1, sheet.max_column + 1):
            if sheet.cell(row=r, column=c).value is not None:
                max_col = max(max_col, c)
                
    # Detect end row (ignore signatures)
    end_row = start_row
    consecutive_empty = 0
    for r in range(start_row, sheet.max_row + 1):
        row_is_part_of_table = False
        
        for c in range(1, max_col + 1):
            cell = sheet.cell(row=r, column=c)
            has_border = cell.border and (cell.border.top.style or cell.border.bottom.style or cell.border.left.style or cell.border.right.style)
            if has_border:
                row_is_part_of_table = True
                break
                
        if row_is_part_of_table:
            end_row = r
            consecutive_empty = 0
        else:
            has_data = False
            for c in range(1, max_col + 1):
                val = sheet.cell(row=r, column=c).value
                if val is not None and str(val).strip() != "":
                    has_data = True
                    break
            
            if has_data:
                val_str = " ".join([str(sheet.cell(row=r, column=c_idx).value) for c_idx in range(1, max_col + 1) if sheet.cell(row=r, column=c_idx).value is not None])
                val_str_up = val_str.upper()
                if "TRƯỞNG BAN" in val_str_up or "CHỮ KÝ" in val_str_up or "ĐẠI TÁ" in val_str_up or "HIỆU TRƯỞNG" in val_str_up or "NGÀY" in val_str_up:
                    break # Stop at signature
                else:
                    end_row = r # Might be table data without borders
                    consecutive_empty = 0
            else:
                consecutive_empty += 1
                if consecutive_empty > 3:
                    break # 3 empty rows probably mean end of table
    
    # Also expand max_col if a merged cell containing data extends further within table bounds
    for (r, c), info in merged_cells_info.items():
        if start_row <= r <= end_row and info['row_span'] > 0: 
            cell = sheet.cell(row=r, column=c)
            if cell.value is not None:
                max_col = max(max_col, c + info['col_span'] - 1)

    # Store cell data (value, rowspan, colspan)
    data = []
    for r in range(start_row, end_row + 1):
        row_data = []
        for c in range(1, max_col + 1):
            cell = sheet.cell(row=r, column=c)
            
            merge_info = merged_cells_info.get((r, c))
            if merge_info:
                row_span = merge_info['row_span']
                col_span = merge_info['col_span']
            else:
                row_span = 1
                col_span = 1
            
            # Truncate rowspan and colspan if they exceed end_row and max_col
            if row_span > 0:
                if r + row_span - 1 > end_row:
                    row_span = end_row - r + 1
                if c + col_span - 1 > max_col:
                    col_span = max_col - c + 1
            
            # Append cell data if it's not a secondary cell in a merge
            if row_span != 0:
                # ensure cell value is at least string if not none
                cell_val = "" if cell.value is None else str(cell.value).strip()
                
                font_color = None
                is_bold = False
                try:
                    if cell.font:
                        is_bold = bool(cell.font.b)
                        if cell.font.color and hasattr(cell.font.color, 'rgb') and cell.font.color.rgb:
                            rgb_val = str(cell.font.color.rgb)
                            if rgb_val.startswith('FF') and len(rgb_val) == 8:
                                font_color = '#' + rgb_val[2:]
                            elif len(rgb_val) == 6 and rgb_val != '000000':
                                font_color = '#' + rgb_val
                except Exception:
                    pass

                row_data.append({
                    "value": cell_val,
                    "rowspan": row_span,
                    "colspan": col_span,
                    "is_bold": is_bold,
                    "font_color": font_color
                })

        data.append(row_data)

    return data

def _safe_cell_value(sheet, r, c):
    if r < sheet.nrows and c < sheet.row_len(r):
        return sheet.cell_value(r, c)
    return None

def _safe_cell_xf_index(sheet, r, c):
    if r < sheet.nrows and c < sheet.row_len(r):
        return sheet.cell_xf_index(r, c)
    return None

def _safe_cell_type(sheet, r, c):
    if r < sheet.nrows and c < sheet.row_len(r):
        return sheet.cell_type(r, c)
    return xlrd.XL_CELL_EMPTY

def _parse_xls(file_path):
    workbook = xlrd.open_workbook(file_path, formatting_info=True)
    sheet = workbook.sheet_by_index(0)

    # Build merged cells info
    merged_cells_info = {}
    for crange in sheet.merged_cells:
        rlo, rhi, clo, chi = crange
        for r in range(rlo, rhi):
            for c in range(clo, chi):
                if r == rlo and c == clo:
                    merged_cells_info[(r+1, c+1)] = {
                        'row_span': rhi - rlo,
                        'col_span': chi - clo
                    }
                else:
                    merged_cells_info[(r+1, c+1)] = {
                        'row_span': 0,
                        'col_span': 0
                    }

    # Detect start row
    start_row = 1
    for r in range(min(sheet.nrows, 50)):
        found_header = False
        for c in range(min(sheet.ncols, 20)):
            val = _safe_cell_value(sheet, r, c)
            if isinstance(val, str) and val.strip().upper() == 'TT':
                start_row = r + 1
                found_header = True
                break
        if found_header:
            break

    # Detect actual max column
    max_col = 1
    for c in range(sheet.ncols):
        val = _safe_cell_value(sheet, start_row - 1, c)
        if val is not None and str(val).strip() != "":
            max_col = max(max_col, c + 1)

    for (r, c), info in merged_cells_info.items():
        if r == start_row and info['row_span'] > 0:
            max_col = max(max_col, c + info['col_span'] - 1)

    for r in range(start_row, min(start_row + 3, sheet.nrows + 1)):
        for c in range(sheet.ncols):
            val = _safe_cell_value(sheet, r - 1, c)
            if val is not None and str(val).strip() != "":
                max_col = max(max_col, c + 1)

    # Detect end row
    end_row = start_row
    consecutive_empty = 0
    for r in range(start_row, sheet.nrows + 1):
        row_is_part_of_table = False
        
        for c in range(1, max_col + 1):
            xf_idx = _safe_cell_xf_index(sheet, r - 1, c - 1)
            if xf_idx is not None:
                xf = workbook.xf_list[xf_idx]
                # Check for borders
                if xf.border.top_line_style > 0 or xf.border.bottom_line_style > 0 or xf.border.left_line_style > 0 or xf.border.right_line_style > 0:
                    row_is_part_of_table = True
                    break

        if row_is_part_of_table:
            end_row = r
            consecutive_empty = 0
        else:
            has_data = False
            for c in range(1, max_col + 1):
                val = _safe_cell_value(sheet, r - 1, c - 1)
                if val is not None and str(val).strip() != "":
                    has_data = True
                    break
            
            if has_data:
                texts = []
                for c_idx in range(1, max_col + 1):
                    v = _safe_cell_value(sheet, r - 1, c_idx - 1)
                    if v is not None:
                        texts.append(str(v))
                val_str = " ".join(texts).upper()
                if "TRƯỞNG BAN" in val_str or "CHỮ KÝ" in val_str or "ĐẠI TÁ" in val_str or "HIỆU TRƯỞNG" in val_str or "NGÀY" in val_str:
                    break
                else:
                    end_row = r
                    consecutive_empty = 0
            else:
                consecutive_empty += 1
                if consecutive_empty > 3:
                    break
                    
    for (r, c), info in merged_cells_info.items():
        if start_row <= r <= end_row and info['row_span'] > 0: 
            val = _safe_cell_value(sheet, r - 1, c - 1)
            if val is not None and str(val).strip() != "":
                max_col = max(max_col, c + info['col_span'] - 1)

    data = []
    for r in range(start_row, end_row + 1):
        row_data = []
        for c in range(1, max_col + 1):
            merge_info = merged_cells_info.get((r, c))
            if merge_info:
                row_span = merge_info['row_span']
                col_span = merge_info['col_span']
            else:
                row_span = 1
                col_span = 1
            
            if row_span > 0:
                if r + row_span - 1 > end_row:
                    row_span = end_row - r + 1
                if c + col_span - 1 > max_col:
                    col_span = max_col - c + 1
            
            if row_span != 0:
                cell_val_str = ""
                font_color = None
                is_bold = False
                
                cell_val = _safe_cell_value(sheet, r - 1, c - 1)
                if cell_val is not None:
                    cell_type = _safe_cell_type(sheet, r - 1, c - 1)
                    
                    if cell_type == xlrd.XL_CELL_DATE:
                        try:
                            dt = xlrd.xldate.xldate_as_datetime(cell_val, workbook.datemode)
                            if dt.hour == 0 and dt.minute == 0 and dt.second == 0:
                                cell_val_str = dt.strftime("%Y-%m-%d")
                            else:
                                cell_val_str = dt.strftime("%Y-%m-%d %H:%M:%S")
                        except:
                            cell_val_str = str(cell_val).strip()
                    elif cell_type == xlrd.XL_CELL_NUMBER:
                        if int(cell_val) == cell_val:
                            cell_val_str = str(int(cell_val))
                        else:
                            cell_val_str = str(cell_val).strip()
                    else:
                        cell_val_str = "" if cell_val is None else str(cell_val).strip()
                        
                    try:
                        xf_idx = _safe_cell_xf_index(sheet, r - 1, c - 1)
                        if xf_idx is not None:
                            xf = workbook.xf_list[xf_idx]
                            font = workbook.font_list[xf.font_index]
                            is_bold = font.weight >= 700
                            
                            color_idx = font.colour_index
                            if color_idx is not None and color_idx != 8 and color_idx != 32767:
                                rgb = workbook.colour_map.get(color_idx)
                                if rgb:
                                    font_color = '#{:02x}{:02x}{:02x}'.format(rgb[0], rgb[1], rgb[2])
                    except Exception:
                        pass

                row_data.append({
                    "value": cell_val_str,
                    "rowspan": row_span,
                    "colspan": col_span,
                    "is_bold": is_bold,
                    "font_color": font_color
                })

        data.append(row_data)

    return data

@app.route('/api/subject_info', methods=['GET'])
def get_subject_info():
    chapter_code = request.args.get('chapter_code')
    if not chapter_code:
        return jsonify({"error": "Thiếu mã bài học (chapter_code) đính kèm."}), 400
        
    try:
        connection = get_db_connection()
        with connection.cursor() as cursor:
            # Join chapter and subject tables to get ChapterName and SubjectName
            sql = """
                SELECT c.ChapterName, s.SubjectName 
                FROM chapter c 
                JOIN subject s ON c.SubjectID = s.ID 
                WHERE c.ChapterCode = %s
            """
            cursor.execute(sql, (chapter_code,))
            result = cursor.fetchone()
            
            if result:
                return jsonify(result), 200
            else:
                return jsonify({"error": f"Không tìm thấy môn học/bài học cho mã '{chapter_code}'."}), 404
    except Exception as e:
        app.logger.error(f"Database error: {e}")
        return jsonify({"error": f"Lỗi cơ sở dữ liệu: {str(e)}"}), 500
    finally:
        if 'connection' in locals() and connection.open:
            connection.close()

@app.route('/api/schedule')
def get_schedule():
    week = request.args.get('week')
    year = request.args.get('year')
    try:
        # Construct the path to the Excel file relative to the script's location
        script_dir = os.path.dirname(__file__)  # a.k.a. 'backend' folder
        project_root = os.path.dirname(script_dir) # a.k.a. 'calendar' folder
        
        file_path = None
        if week:
            safe_week = week.replace('/', '_').replace('\\', '_')
            path_prefix = safe_week
            if year:
                safe_year = year.replace('/', '_').replace('\\', '_')
                path_prefix = f"{safe_year}_{safe_week}"
                
            path_xlsx = os.path.join(app.config['UPLOAD_FOLDER'], f"{path_prefix}.xlsx")
            path_xls = os.path.join(app.config['UPLOAD_FOLDER'], f"{path_prefix}.xls")
            if os.path.exists(path_xlsx):
                file_path = path_xlsx
            elif os.path.exists(path_xls):
                file_path = path_xls
            else:
                return jsonify({"error": f"Không tìm thấy dữ liệu lịch cho '{week}' năm học '{year or 'Mặc định'}'. Vui lòng tải lên file lịch."}), 404
        else:
            path_xlsx = os.path.join(project_root, 'sample.xlsx')
            path_xls = os.path.join(project_root, 'sample.xls')
            if os.path.exists(path_xls): # prioritize xls as we are testing it, or xlsx. Let's do xlsx first.
                file_path = path_xls
            elif os.path.exists(path_xlsx):
                file_path = path_xlsx
            else:
                return jsonify({"error": "No sample file found."}), 404

        if file_path.lower().endswith('.xls'):
            data = _parse_xls(file_path)
        else:
            data = _parse_xlsx(file_path)

        return jsonify(data)
    except FileNotFoundError:
        return jsonify({"error": "The file 'sample.xls' was not found."}), 404
    except Exception as e:
        # It's good practice to log the actual exception
        app.logger.error(f"An error occurred: {e}")
        return jsonify({"error": str(e)}), 500

if __name__ == '__main__':
    # Running on port 5001 to avoid potential conflicts with other services
    app.run(debug=True, port=5001)
