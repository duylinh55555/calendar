import openpyxl
import io
import sys

# Change standard output to utf-8 encoding to handle Vietnamese chars
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

file_path = 'd:\\VS Code Projects\\calendar\\sample.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb.active

print("First 15 rows:")
for r in range(1, min(ws.max_row + 1, 16)):
    row_vals = [str(ws.cell(row=r, column=c).value) if ws.cell(row=r, column=c).value is not None else '' for c in range(1, min(ws.max_column + 1, 20))]
    # Print only non-empty columns to avoid too much text
    row_vals_trimmed = [v for v in row_vals if v.strip() != '']
    print(f"Row {r}: {row_vals_trimmed}")

print("\nLast 15 rows:")
start_last = max(1, ws.max_row - 15)
for r in range(start_last, ws.max_row + 1):
    row_vals = [str(ws.cell(row=r, column=c).value) if ws.cell(row=r, column=c).value is not None else '' for c in range(1, min(ws.max_column + 1, 20))]
    row_vals_trimmed = [v for v in row_vals if v.strip() != '']
    print(f"Row {r}: {row_vals_trimmed}")
